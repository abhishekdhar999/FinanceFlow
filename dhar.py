import tkinter as tk
from tkinter import ttk
import pandas as pd
import datetime as dt
from fyers_apiv3 import fyersModel  # Import Fyers API model
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import product

# Default values for the trading strategy parameters
DEFAULT_VALUES = {
    "Trading Symbol": "NSE:NIFTYBANK-INDEX",
    "Start Time": "09:16",
    "End Time": "15:16",
    "Target 1": 100,
    "Stop Loss": 100,
    "Divider": 2000,
    "Lot": 5,
    "Target 1 Lot": 2,
    "Interval": 1,
    "Buffer": 0.5,
    "Donchian Channel": 20,
    "TSL": 1
}

# Range of values for optimization
PARAM_RANGES = {
    "Divider": [700, 750, 800, 1000, 1200, 1500, 1800, 2000],
    "Buffer": [0.2, 0.3, 0.4, 0.5, 0.7, 0.8, 0.9],
    "Donchian Channel": [15, 17, 20, 23, 25, 27, 30],
    "TSL": [1, 25, 30, 50, 100],
    "Stop Loss": [70, 100, 120, 150, 200],
    "Target 1 Lot": [1, 2, 3, 4, 5],
    "Target 1": [70, 100, 120, 150, 200]
}
# Fyers API credentials
client_id = "client id of fyres"
secret_key = "secret key of fyres "
redirect_uri = "fyres redirecting uri"
response_type = "code"
grant_type = "authorization_code"

# Initialize session
session = fyersModel.SessionModel(
    client_id=client_id,
    secret_key=secret_key,
    redirect_uri=redirect_uri,
    response_type=response_type,
    grant_type=grant_type
)

# Generate the auth code using the session model
response = session.generate_authcode()
print(response)
auth_code = input("Enter the auth code: ")

# Set the authorization code in the session object
session.set_token(auth_code)
response = session.generate_token()
print(response)
access_token = response["access_token"]
# Initialize Fyers model (Replace with actual credentials)
fyers = fyersModel.FyersModel(client_id="fyres client id", token=access_token)

# Additional variables for tracking positions and metrics
positions = []
metrics = {
    "No of Positions": 0,
    "Winning trade": 0,
    "Winning %": 0,
    "max profit": 0,
    "overall profit": 0,
    "max loss": 0,
    "winning streak": 0,
    "loss streak": 0
}

def fetch_fyers_data():
    try:
        print("Fetching historical data from Fyers API...")
        now = dt.datetime.now()
        end_time = now
        start_time = end_time - dt.timedelta(days=100)

        start_str = start_time.strftime('%Y-%m-%d')
        end_str = end_time.strftime('%Y-%m-%d')

        data = {
            "symbol": DEFAULT_VALUES["Trading Symbol"],
            "resolution": "1",
            "date_format": "1",
            "range_from": start_str,
            "range_to": end_str,
            "cont_flag": "1"
        }

        response = fyers.history(data=data)

        if response['code'] != 200:
            print(f"Error fetching data: {response}")
            return pd.DataFrame()

        candles = response['candles']
        df = pd.DataFrame(candles)

        if df.empty:
            print("Received empty data from the API.")
            return pd.DataFrame()

        df.columns = ['date', 'open', 'high', 'low', 'close', 'volume']
        df['date'] = pd.to_datetime(df['date'], unit='s')
        df['date'] = df['date'].dt.tz_localize('UTC').dt.tz_convert('Asia/Kolkata')
        df['date'] = df['date'].dt.tz_localize(None)
        df.set_index('date', inplace=True)

        print("Historical data fetched successfully.")
        return df

    except Exception as e:
        print(f"Exception in fetch_fyers_data(): {str(e)}")
        return pd.DataFrame()

def generate_renko(df, divider):
    try:
        print("Generating Renko bricks with divider:", divider)
        renko_bricks = []
        renko_dates = []
        renko_sizes = []

        if divider == 0:
            print("Divider should not be zero. Please check your parameters.")
            return pd.DataFrame()

        grouped = df.groupby(df.index.date)

        for date, group in grouped:
            opening_price = group['open'].iloc[0]
            brick_size = round(opening_price / divider)
            previous_close = opening_price

            for i in range(len(group)):
                price_move = group['close'].iloc[i] - previous_close

                if abs(price_move) >= brick_size:
                    if price_move > 0:
                        renko_bricks.append(previous_close + brick_size)
                        previous_close += brick_size
                    else:
                        renko_bricks.append(previous_close - brick_size)
                        previous_close -= brick_size

                    renko_dates.append(group.index[i])
                    renko_sizes.append(brick_size)
                    previous_close = group['close'].iloc[i]

        renko_df = pd.DataFrame({'date': renko_dates, 'high': renko_bricks, 'low': renko_bricks,
                                 'close': renko_bricks, 'open': [opening_price] * len(renko_bricks), 'brick_size': renko_sizes})
        renko_df.set_index('date', inplace=True)

        print("Renko bricks generated successfully.")
        return renko_df

    except Exception as e:
        print(f"Exception in generate_renko(): {str(e)}")
        return pd.DataFrame()

def calculate_donchian_channels(df, donchian_period=20):
    try:
        df['donchian_high'] = df['close'].rolling(window=donchian_period).max()
        df['donchian_low'] = df['close'].rolling(window=donchian_period).min()
        df.dropna(inplace=True)
        return df

    except Exception as e:
        print(f"Exception in calculate_donchian_channels(): {str(e)}")
        return pd.DataFrame()

def generate_signals(df, donchian_period=20, buffer=0.5, lot_size=75, target_1=100, tsl=1, stop_loss=100, target_1_lot=2):
    try:
        df = calculate_donchian_channels(df, donchian_period)

        if df.empty:
            raise ValueError("Dataframe is empty after calculating Donchian channels.")

        signals = pd.DataFrame(index=df.index)
        signals['Buy CE'] = 0
        signals['Buy PE'] = 0
        signals['Position'] = 0
        signals['Entry Price'] = 0.0
        signals['Exit Price'] = 0.0
        signals['Profit/Loss'] = 0.0
        signals['Reason'] = ''

        brick_size = df['brick_size'].iloc[0]

        entry_price = 0.0
        current_position = 0
        cumulative_move = 0.0
        remaining_lots = lot_size
        target_1_hit = False
        combined_profit_loss = 0.0
        trailing_stop_loss = 0.0

        for i in range(1, len(df)):
            price_move = df['close'].iloc[i] - df['close'].iloc[i - 1]
            cumulative_move += price_move

            # Check if we are within trading hours
            if df.index[i].time() >= dt.time(9, 16) and df.index[i].time() <= dt.time(15, 16):
                # No active position
                if current_position == 0 and not target_1_hit:
                    # Check for Buy CE signal
                    if cumulative_move >= brick_size + buffer * brick_size and df['close'].iloc[i] > df['donchian_high'].iloc[i - 1]:
                        signals.at[df.index[i], 'Buy CE'] = 1
                        signals.at[df.index[i], 'Position'] = 1
                        signals.at[df.index[i], 'Entry Price'] = df['close'].iloc[i]
                        signals.at[df.index[i], 'Reason'] = 'Buy CE signal triggered'
                        entry_price = df['close'].iloc[i]
                        current_position = 1
                        cumulative_move = 0.0
                        trailing_stop_loss = entry_price - stop_loss
                    # Check for Buy PE signal
                    elif cumulative_move <= -(brick_size + buffer * brick_size) and df['close'].iloc[i] < df['donchian_low'].iloc[i - 1]:
                        signals.at[df.index[i], 'Buy PE'] = 1
                        signals.at[df.index[i], 'Position'] = -1
                        signals.at[df.index[i], 'Entry Price'] = df['close'].iloc[i]
                        signals.at[df.index[i], 'Reason'] = 'Buy PE signal triggered'
                        entry_price = df['close'].iloc[i]
                        current_position = -1
                        cumulative_move = 0.0
                        trailing_stop_loss = entry_price + stop_loss

                if current_position == 1:
                    trailing_stop_loss = max(trailing_stop_loss, df['close'].iloc[i] - tsl)
                    if df['close'].iloc[i] >= entry_price + target_1 and not target_1_hit and df['close'].iloc[i] > df['donchian_high'].iloc[i - 1]:
                        profit_loss = (df['close'].iloc[i] - entry_price) * target_1_lot * 15
                        combined_profit_loss = profit_loss
                        signals.at[df.index[i], 'Exit Price'] = df['close'].iloc[i]
                        signals.at[df.index[i], 'Reason'] = f'Target 1 hit - {profit_loss:.2f}'
                        remaining_lots -= target_1_lot * 15
                        target_1_hit = True
                    elif df['close'].iloc[i] <= trailing_stop_loss or df['close'].iloc[i] < df['donchian_low'].iloc[i - 1]:
                        signals.at[df.index[i], 'Exit Price'] = df['close'].iloc[i]
                        signals.at[df.index[i], 'Profit/Loss'] = (df['close'].iloc[i] - entry_price) * lot_size
                        signals.at[df.index[i], 'Reason'] = 'Stop loss hit'
                        current_position = 0
                        entry_price = 0.0
                        target_1_hit = False
                        remaining_lots = lot_size
                    elif target_1_hit and df['close'].iloc[i] >= trailing_stop_loss:
                        combined_profit_loss += (df['close'].iloc[i] - entry_price) * remaining_lots
                        signals.at[df.index[i], 'Exit Price'] = df['close'].iloc[i]
                        signals.at[df.index[i], 'Profit/Loss'] = combined_profit_loss
                        signals.at[df.index[i], 'Reason'] = 'Trailing stop loss hit'
                        current_position = 0
                        entry_price = 0.0
                        target_1_hit = False
                        remaining_lots = lot_size
                        combined_profit_loss = 0.0

                elif current_position == -1:
                    trailing_stop_loss = min(trailing_stop_loss, df['close'].iloc[i] + tsl)
                    if df['close'].iloc[i] <= entry_price - target_1 and not target_1_hit and df['close'].iloc[i] < df['donchian_low'].iloc[i - 1]:
                        profit_loss = (entry_price - df['close'].iloc[i]) * target_1_lot * 15
                        combined_profit_loss = profit_loss
                        signals.at[df.index[i], 'Exit Price'] = df['close'].iloc[i]
                        signals.at[df.index[i], 'Reason'] = f'Target 1 hit - {profit_loss:.2f}'
                        remaining_lots -= target_1_lot * 15
                        target_1_hit = True
                    elif df['close'].iloc[i] >= trailing_stop_loss or df['close'].iloc[i] > df['donchian_high'].iloc[i - 1]:
                        signals.at[df.index[i], 'Exit Price'] = df['close'].iloc[i]
                        signals.at[df.index[i], 'Profit/Loss'] = (entry_price - df['close'].iloc[i]) * lot_size
                        signals.at[df.index[i], 'Reason'] = 'Stop loss hit'
                        current_position = 0
                        entry_price = 0.0
                        target_1_hit = False
                        remaining_lots = lot_size
                    elif target_1_hit and df['close'].iloc[i] <= trailing_stop_loss:
                        combined_profit_loss += (entry_price - df['close'].iloc[i]) * remaining_lots
                        signals.at[df.index[i], 'Exit Price'] = df['close'].iloc[i]
                        signals.at[df.index[i], 'Profit/Loss'] = combined_profit_loss
                        signals.at[df.index[i], 'Reason'] = 'Trailing stop loss hit'
                        current_position = 0
                        entry_price = 0.0
                        target_1_hit = False
                        remaining_lots = lot_size
                        combined_profit_loss = 0.0

            # End of day exit based on last Renko Brick
            if i == len(df) - 1 or df.index[i].date() != df.index[i + 1].date():
                if current_position != 0:
                    profit_loss = (df['close'].iloc[i] - entry_price) * remaining_lots * current_position
                    combined_profit_loss += profit_loss
                    signals.at[df.index[i], 'Exit Price'] = df['close'].iloc[i]
                    signals.at[df.index[i], 'Profit/Loss'] = combined_profit_loss
                    signals.at[df.index[i], 'Reason'] = 'End of day exit'
                    current_position = 0
                    entry_price = 0.0
                    target_1_hit = False
                    remaining_lots = lot_size
                    combined_profit_loss = 0.0

        return signals

    except Exception as e:
        print(f"Exception in generate_signals(): {str(e)}")
        return pd.DataFrame()

def update_metrics(signals):
    try:
        metrics['No of Positions'] = len(signals[signals['Position'] != 0])
        metrics['Winning trade'] = len(signals[signals['Profit/Loss'] > 0])
        metrics['Winning %'] = (metrics['Winning trade'] / metrics['No of Positions']) * 100 if metrics['No of Positions'] > 0 else 0
        metrics['max profit'] = signals['Profit/Loss'].max()
        metrics['overall profit'] = signals['Profit/Loss'].sum()
        metrics['max loss'] = signals['Profit/Loss'].min()

        winning_streak = 0
        loss_streak = 0
        current_winning_streak = 0
        current_loss_streak = 0

        for profit_loss in signals['Profit/Loss']:
            if profit_loss > 0:
                current_winning_streak += 1
                current_loss_streak = 0
                winning_streak = max(winning_streak, current_winning_streak)
            elif profit_loss < 0:
                current_loss_streak += 1
                current_winning_streak = 0
                loss_streak = max(loss_streak, current_loss_streak)

        metrics['winning streak'] = winning_streak
        metrics['loss streak'] = loss_streak

        return metrics

    except Exception as e:
        print(f"Exception in update_metrics(): {str(e)}")
        return {}
    
def save_to_excel(signals, metrics, renko_df, all_metrics):
    try:
        print("Saving data to Excel...")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Signals"

        for r in dataframe_to_rows(signals, index=True, header=True):
            ws1.append(r)

        ws2 = wb.create_sheet(title="Metrics")
        for key, value in metrics.items():
            ws2.append([key, value])

        ws3 = wb.create_sheet(title="Renko Bricks")
        for r in dataframe_to_rows(renko_df, index=True, header=True):
            ws3.append(r)

        ws4 = wb.create_sheet(title="All Metrics")
        for r in dataframe_to_rows(all_metrics, index=False, header=True):
            ws4.append(r)

        wb.save("trading_signals.xlsx")
        print("Data saved to trading_signals.xlsx successfully.")

    except Exception as e:
        print(f"Exception in save_to_excel(): {str(e)}")

def run_strategy():
    print("Running trading strategy...")
    df = fetch_fyers_data()

    if df.empty:
        print("No data fetched. Exiting the strategy.")
        return

    renko_df = generate_renko(df, DEFAULT_VALUES['Divider'])
    if renko_df.empty:
        print("No Renko data generated. Exiting the strategy.")
        return

    best_params = None
    best_winning_percentage = 0
    total_combinations = len(list(product(*PARAM_RANGES.values())))
    processed_combinations = 0

    all_metrics = []
    param_combinations = product(*PARAM_RANGES.values())
    for params in param_combinations:
        print(f"Testing parameters: Divider={params[0]}, Buffer={params[1]}, Donchian Channel={params[2]}, TSL={params[3]}, Stop Loss={params[4]}, Target 1 Lot={params[5]}, Target 1={params[6]}")
        processed_combinations += 1
        print(f"Processed {processed_combinations}/{total_combinations} combinations. Pending: {total_combinations - processed_combinations}")
        divider, buffer, donchian_channel, tsl, stoploss, target1lot, target1   = params

        renko_df = generate_renko(df, divider)
        if renko_df.empty:
            continue

        signals = generate_signals(renko_df, donchian_period=donchian_channel, buffer=buffer, tsl=tsl, stop_loss=stoploss, target_1_lot=target1lot, target_1=target1 )
        if signals.empty:
            continue

        metrics = update_metrics(signals)
        print(f"Metrics for current parameters: {metrics}")

        # Collecting metrics for all combinations
        all_metrics.append({
            "Divider": divider,
            "Buffer": buffer,
            "Donchian Channel": donchian_channel,
            "TSL": tsl,
            "Stop Loss": stoploss,
            "Target 1 Lot": target1lot,
            "Target 1": target1,
            "No of Positions": metrics['No of Positions'],
            "Winning trade": metrics['Winning trade'],
            "Winning %": metrics['Winning %'],
            "max profit": metrics['max profit'],
            "overall profit": metrics['overall profit'],
            "max loss": metrics['max loss'],
            "winning streak": metrics['winning streak'],
            "loss streak": metrics['loss streak']
        })

        if metrics['Winning %'] > best_winning_percentage:
            best_winning_percentage = metrics['Winning %']
            best_params = {
                "Divider": divider,
                "Buffer": buffer,
                "Donchian Channel": donchian_channel,
                "TSL": tsl,
                "Stop Loss": stoploss,
                "Target 1 Lot": target1lot,
                "Target 1": target1
            }

    print(f"Best parameters: {best_params}")
    print(f"Best winning percentage: {best_winning_percentage}")

    if best_params:
        # Re-run with the best parameters
        renko_df = generate_renko(df, best_params['Divider'])
        signals = generate_signals(renko_df, donchian_period=best_params['Donchian Channel'], buffer=best_params['Buffer'], tsl=best_params['TSL'], stop_loss=best_params['Stop Loss'], target_1_lot=best_params['Target 1 Lot'], target_1=best_params['Target 1'])
        metrics = update_metrics(signals)
        
        all_metrics_df = pd.DataFrame(all_metrics)
        save_to_excel(signals, metrics, renko_df, all_metrics_df)
        print("Strategy completed successfully with optimized parameters.")


# GUI code here

root = tk.Tk()
root.title("Trading Strategy Parameters")

# Parameter input widgets here

button = ttk.Button(root, text="Run Strategy", command=run_strategy)
button.pack()

root.mainloop()